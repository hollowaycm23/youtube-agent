# -*- coding: utf-8 -*-
"""RAG (Retrieval-Augmented Generation) system.

This module provides utilities to:
- Load and embed documents from a folder.
- Transcribe audio/video files with Whisper.
- Retrieve the most relevant context for a query using a GPU-accelerated FAISS index.

All heavy model loading is performed lazily and uses CUDA when available.
"""

import os
import json
import time
import logging
import requests
import shutil
from pathlib import Path
from typing import List, Tuple, Optional, Dict, Any

import numpy as np

# FAISS for GPU-accelerated indexing
try:
    import faiss
except ImportError as exc:
    raise ImportError(
        "FAISS not installed. Please install with: pip install faiss-gpu-cu12"
    ) from exc


# Optional dependencies (transformers, torch, document readers)
try:
    from transformers import AutoTokenizer, AutoModel, pipeline
    import torch
except ImportError as exc:
    raise ImportError(
        "Critical dependencies missing: transformers and torch. "
        "Install with: pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cu121 && pip install transformers pypdf docx openpyxl"
    ) from exc

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration constants
# ---------------------------------------------------------------------------
RAG_FOLDER: str = "rag_documents"
RAG_FAISS_INDEX: str = "rag_faiss.index"
RAG_METADATA_FILE: str = "rag_metadata.json"
EMBEDDING_DIM: int = 1024  # Dimension of multilingual-e5-large embeddings
MAX_CONTEXT_TOKENS: int = 4000
SIMILARITY_THRESHOLD: float = 0.7  # configurable similarity cut-off
DEVICE: str = "cuda" if torch.cuda.is_available() else "cpu"

# Embedding model (multilingual-e5-large)
EMBEDDING_MODEL_ID: str = "intfloat/multilingual-e5-large"
EMBEDDING_MODEL: Optional[torch.nn.Module] = None
EMBEDDING_TOKENIZER: Optional[Any] = None

# Whisper model – medium size for Portuguese transcription
WHISPER_MODEL_ID: str = "openai/whisper-medium"
WHISPER_PIPELINE: Optional[Any] = None

# ---------------------------------------------------------------------------
# Helper – configure module-wide logger
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

def check_hf_connection() -> bool:
    """Simple connectivity test to Hugging Face hub."""
    if os.getenv("HF_HUB_OFFLINE"):
        logger.info("HF_HUB_OFFLINE is set – skipping Hugging Face connectivity check.")
        return True
    try:
        resp = requests.get("https://huggingface.co", timeout=5)
        logger.info("Hugging Face connection status: %s", resp.status_code)
        return resp.ok
    except Exception as exc:
        logger.error("Unable to reach Hugging Face: %s", exc)
        return False

# ---------------------------------------------------------------------------
# Model initialization
# ---------------------------------------------------------------------------
def initialize_models() -> Tuple[Optional[torch.nn.Module], Optional[Any], Optional[Any], str]:
    """Load the embedding and Whisper models."""
    logger.info("Initializing models on device %s", DEVICE.upper())
    if not os.getenv("HF_HUB_OFFLINE") and not check_hf_connection():
        logger.error("Cannot connect to Hugging Face. Model loading aborted.")
        return None, None, None, DEVICE

    device = DEVICE
    if device == "cuda" and not torch.cuda.is_available():
        logger.warning("CUDA selected but not available – falling back to CPU")
        device = "cpu"

    # Embedding model
    try:
        tokenizer = AutoTokenizer.from_pretrained(EMBEDDING_MODEL_ID)
        model = AutoModel.from_pretrained(EMBEDDING_MODEL_ID).to(device)
        model.eval()
        logger.info("Embedding model loaded: %s", EMBEDDING_MODEL_ID)
    except Exception as exc:
        logger.error("Failed to load embedding model: %s", exc)
        tokenizer, model = None, None

    # Whisper pipeline
    try:
        pipe = pipeline(
            "automatic-speech-recognition",
            model=WHISPER_MODEL_ID,
            device=0 if device == "cuda" else -1,
        )
        logger.info("Whisper model loaded: %s", WHISPER_MODEL_ID)
    except Exception as exc:
        logger.error("Failed to load Whisper model: %s", exc)
        pipe = None

    return model, tokenizer, pipe, device

# ---------------------------------------------------------------------------
# Text extraction utilities
# ---------------------------------------------------------------------------
def extract_text_from_file(filepath: str) -> str:
    """Extract plain text from supported file types."""
    _, ext = os.path.splitext(filepath)
    ext = ext.lower()
    try:
        if ext in (".txt", ".md", ".py", ".js", ".html", ".css"):
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        if ext == ".pdf":
            reader = PdfReader(filepath)
            return "".join(page.extract_text() or "" for page in reader.pages)
        if ext == ".docx":
            doc = Document(filepath)
            return "\n".join(p.text for p in doc.paragraphs)
        if ext == ".xlsx":
            wb = load_workbook(filepath)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"--- Planilha: {sheet_name} ---\n"
                for row in sheet.iter_rows():
                    text += " ".join(str(cell.value) if cell.value is not None else "" for cell in row) + "\n"
            return text
        logger.warning("Unsupported file extension: %s", ext)
        return ""
    except Exception as exc:
        logger.error("Error reading %s: %s", filepath, exc)
        return f"ERRO ao ler {ext.upper()}: {exc}"

# ---------------------------------------------------------------------------
# Audio / video transcription
# ---------------------------------------------------------------------------
def transcribe_audio_video(filepath: str) -> Optional[str]:
    """Transcribe an audio/video file using Whisper."""
    global WHISPER_PIPELINE
    if WHISPER_PIPELINE is None:
        logger.warning("Whisper pipeline not initialized – transcription disabled")
        return None
    logger.info("Transcribing %s with Whisper", os.path.basename(filepath))
    try:
        result = WHISPER_PIPELINE(filepath)
        transcription = f"--- TRANSCRIÇÃO DE ÁUDIO/VÍDEO: {os.path.basename(filepath)} ---\n{result['text']}"
        cache_path = f"{filepath}.transcribed.txt"
        with open(cache_path, "w", encoding="utf-8") as f:
            f.write(transcription)
        logger.info("Transcription cached at %s", cache_path)
        return transcription
    except Exception as exc:
        logger.error("Transcription failed for %s: %s", filepath, exc)
        return None

# ---------------------------------------------------------------------------
# Embedding utility
# ---------------------------------------------------------------------------
def calculate_embedding(text: str) -> Optional[np.ndarray]:
    """Return a normalized embedding vector for *text*."""
    global EMBEDDING_MODEL, EMBEDDING_TOKENIZER
    if EMBEDDING_MODEL is None or EMBEDDING_TOKENIZER is None:
        logger.warning("Embedding model not initialized. Cannot calculate embedding.")
        return None
    try:
        input_text = f"query: {text}"
        encoded = EMBEDDING_TOKENIZER(
            input_text,
            padding=True,
            truncation=True,
            max_length=512,
            return_tensors="pt",
        ).to(DEVICE)
        with torch.no_grad():
            output = EMBEDDING_MODEL(**encoded)
        emb = output.last_hidden_state[:, 0]
        emb = torch.nn.functional.normalize(emb, p=2, dim=1)
        return emb[0].cpu().numpy()
    except Exception as exc:
        logger.error("Embedding error: %s", exc)
        return None

# ---------------------------------------------------------------------------
# Vector database creation / update
# ---------------------------------------------------------------------------
def create_vector_db() -> bool:
    """Create or update a FAISS index and a metadata JSON file."""
    global EMBEDDING_MODEL, EMBEDDING_TOKENIZER, WHISPER_PIPELINE, DEVICE
    if not os.path.isdir(RAG_FOLDER):
        os.makedirs(RAG_FOLDER, exist_ok=True)
        logger.info("Created RAG folder: %s", RAG_FOLDER)

    logger.info("Building/updating vector database with FAISS")
    EMBEDDING_MODEL, EMBEDDING_TOKENIZER, WHISPER_PIPELINE, DEVICE = initialize_models()
    if EMBEDDING_MODEL is None:
        logger.error("Embedding model could not be loaded – aborting DB creation")
        return False

    metadata: List[Dict[str, Any]] = []
    embeddings: List[np.ndarray] = []
    
    supported_exts = {'.txt', '.md', '.py', '.js', '.html', '.css', '.pdf', '.docx', '.xlsx'}
    media_exts = {'.mp4', '.mp3', '.mov', '.wav'}

    for filename in os.listdir(RAG_FOLDER):
        path = os.path.join(RAG_FOLDER, filename)
        if not os.path.isfile(path):
            continue
        
        logger.info("Processing %s", filename)
        content: Optional[str] = None
        
        _, ext = os.path.splitext(path)
        ext = ext.lower()

        if ext in media_exts:
            content = transcribe_audio_video(path)
        elif ext in supported_exts:
            content = extract_text_from_file(path)
        else:
            logger.warning("Skipping unsupported file type: %s", filename)
            continue
            
        if not content or content.startswith("ERRO ao ler"):
            logger.warning("Skipping %s – content extraction failed", filename)
            continue

        embedding = calculate_embedding(content)
        if embedding is None:
            logger.warning("Embedding failed for %s – skipping", filename)
            continue

        embeddings.append(embedding)
        metadata.append({
            "filepath": path,
            "content": content,
        })
        logger.info("Added/updated %s to be indexed", filename)

    if not embeddings:
        logger.warning("No documents to index. Aborting.")
        return False

    # Create and populate FAISS index
    logger.info("Creating FAISS index...")
    index = faiss.IndexFlatL2(EMBEDDING_DIM)
    if DEVICE == "cuda":
        logger.info("Moving FAISS index to GPU...")
        res = faiss.StandardGpuResources()
        index = faiss.index_cpu_to_gpu(res, 0, index)
    
    index.add(np.array(embeddings).astype('float32'))
    
    # Save index and metadata
    if DEVICE == "cuda":
        logger.info("Moving FAISS index back to CPU for saving...")
        index_to_save = faiss.index_gpu_to_cpu(index)
    else:
        index_to_save = index
        
    faiss.write_index(index_to_save, RAG_FAISS_INDEX)
    with open(RAG_METADATA_FILE, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)
        
    logger.info("FAISS index and metadata written with %d documents", len(metadata))
    return True

# ---------------------------------------------------------------------------
# Retrieval
# ---------------------------------------------------------------------------
def retrieve_context(index: faiss.Index, metadata: List[Dict[str, Any]], query_text: str, k: int = 5) -> str:
    """Return the most relevant document snippets for *query_text* using FAISS."""
    query_emb = calculate_embedding(query_text)
    if query_emb is None:
        logger.error("Failed to embed query – aborting retrieval")
        return ""
    
    query_vec = np.expand_dims(query_emb, axis=0).astype('float32')

    # FAISS search returns distances (D) and indices (I)
    distances, indices = index.search(query_vec, k)
    
    context = ""
    total_len = 0
    
    for i, idx in enumerate(indices[0]):
        # The distance from IndexFlatL2 is squared L2, not cosine similarity.
        # For normalized vectors, L2_dist^2 = 2 - 2 * cos_sim. So, cos_sim = 1 - (L2_dist^2 / 2).
        similarity = 1 - (distances[0][i] / 2)
        
        if similarity < SIMILARITY_THRESHOLD or idx < 0:
            continue
            
        item = metadata[idx]
        content = item["content"]
        fname = os.path.basename(item["filepath"])
        
        chunk = f"--- Documento de Contexto ({fname}, Sim: {similarity:.2f}) ---\n{content[:2000]}\n"
        
        if total_len + len(chunk) > MAX_CONTEXT_TOKENS:
            break
        context += chunk
        total_len += len(chunk)
        
    return context.strip()

# ---------------------------------------------------------------------------
# Public function for external modules (e.g., youtube_seo_automation)
# ---------------------------------------------------------------------------
def get_context(query: str) -> str:
    """Convenient wrapper: load DB and index, then return context for a query."""
    global EMBEDDING_MODEL, EMBEDDING_TOKENIZER, WHISPER_PIPELINE, DEVICE
    if EMBEDDING_MODEL is None:
        EMBEDDING_MODEL, EMBEDDING_TOKENIZER, WHISPER_PIPELINE, DEVICE = initialize_models()

    if not os.path.isfile(RAG_FAISS_INDEX) or not os.path.isfile(RAG_METADATA_FILE):
        logger.info("FAISS index or metadata not found – building now")
        if not create_vector_db():
            return "Erro: Não foi possível criar o banco de dados de vetores."
    
    try:
        logger.info("Loading FAISS index and metadata...")
        index = faiss.read_index(RAG_FAISS_INDEX)
        if DEVICE == "cuda":
             res = faiss.StandardGpuResources()
             index = faiss.index_cpu_to_gpu(res, 0, index)

        with open(RAG_METADATA_FILE, "r", encoding="utf-8") as f:
            metadata = json.load(f)
            
    except Exception as exc:
        logger.error("Failed to load FAISS index/metadata: %s. Rebuilding.", exc)
        if not create_vector_db():
            return "Erro: Falha ao recriar o banco de dados de vetores."
        # Try loading again after rebuild
        index = faiss.read_index(RAG_FAISS_INDEX)
        with open(RAG_METADATA_FILE, "r", encoding="utf-8") as f:
            metadata = json.load(f)

    return retrieve_context(index, metadata, query)

# ---------------------------------------------------------------------------
# Optional FastAPI server (only if FastAPI is installed)
# ---------------------------------------------------------------------------
try:
    from fastapi import FastAPI, HTTPException
    from pydantic import BaseModel
    app = FastAPI()

    class QueryRequest(BaseModel):
        query: str

    @app.post("/query")
    async def query_endpoint(req: QueryRequest):
        try:
            context = get_context(req.query)
            if not context:
                raise HTTPException(status_code=404, detail="No relevant context found")
            return {"context": context}
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc))
except ImportError:
    logger.info("FastAPI not installed – API server disabled")

# ---------------------------------------------------------------------------
# CLI interface
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="RAG demo / utility")
    parser.add_argument("--query", type=str, help="Query string to retrieve context")
    parser.add_argument("--rebuild", action="store_true", help="Force rebuild of the FAISS index")
    parser.add_argument("--serve", action="store_true", help="Run FastAPI server (requires fastapi & uvicorn)")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

    if args.serve:
        try:
            import uvicorn
            uvicorn.run(app, host="0.0.0.0", port=8000)
        except ImportError:
            logger.error("uvicorn not installed – cannot start server")
    
    elif args.rebuild:
        logger.info("Forcing rebuild of vector database...")
        create_vector_db()

    elif args.query:
        result = get_context(args.query)
        if result:
            print("--- Retrieved Context ---")
            print(result)
        else:
            print("No relevant context found.")
    else:
        # Default demo behavior: build if not present, then run a test query
        if not os.path.isfile(RAG_FAISS_INDEX):
            logger.info("Running initial database build...")
            create_vector_db()
        
        logger.info("Running RAG demo query")
        test_q = "o que é o Antigravity?"
        logger.info("Query: %s", test_q)
        result = get_context(test_q)
        logger.info("Retrieved context:\n%s", result if result else "Nenhum contexto encontrado")
